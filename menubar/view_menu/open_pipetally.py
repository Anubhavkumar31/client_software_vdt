import sys
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout,
    QLineEdit, QTableView, QFrame, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QMessageBox, QRadioButton, QGridLayout, QSpinBox, QDialog, QProgressDialog,
    QProgressBar
)

from PyQt6.QtCore import QThread, pyqtSignal
from PyQt6.QtGui import QStandardItemModel, QStandardItem, QFont
from PyQt6.QtCore import Qt, QSortFilterProxyModel, QTimer

from menubar.view_menu.apps.pipetallyApp.erf_pipeline import launch_erf_batch
from menubar.view_menu.apps.pipetallyApp.severity_pipeline import launch_severity

# EXCEL_PATH = r"C:\Users\admin\Downloads\Pipe_Tally_8inch (1).xlsx"


# ================= ENTERPRISE TOOL BAR =================

class PipeTallyToolPanel(QFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.setFixedHeight(56)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(18, 8, 18, 8)
        layout.setSpacing(12)

        title = QLabel("PIPE TALLY")
        title.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        title.setStyleSheet("color:#1e293b;")
        layout.addWidget(title)
        layout.addSpacing(10)

        layout.addWidget(self._vline())

        self.btn_search = self._tool_btn("Search")

        self.btn_erf  = self._tool_btn("erf calc")
        self.btn_severity = self._tool_btn("Severity calc")
        self.btn_export = self._tool_btn("Export Excel")
        self.btn_more   = self._tool_btn("More")


        for b in (self.btn_search,self.btn_erf, self.btn_severity, self.btn_export,  self.btn_more):
            layout.addWidget(b)

        layout.addStretch()

        self.setStyleSheet("""
            QFrame {
                background:qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 #ffffff, stop:1 #f1f5f9);
                border-bottom:1px solid #cbd5f5;
            }
        """)

        self.btn_export.clicked.connect(self.export_excel)
        try:
            self.btn_erf.clicked.connect(lambda: launch_erf_batch(self.parent, self.parent.parent().pipetally_dir))
            self.btn_severity.clicked.connect(lambda: launch_severity(self.parent))

        except Exception as e:
            print(f" error {e}")
        self.btn_search.clicked.connect(self.parent.toggle_search_bar)

    def _tool_btn(self, text):
        b = QPushButton(text)
        b.setFixedHeight(34)
        b.setCursor(Qt.CursorShape.PointingHandCursor)
        b.setStyleSheet("""
            QPushButton {
                background:white;
                border:1px solid #cbd5f5;
                border-radius:8px;
                padding:6px 18px;
                color:#1e293b;
                font-weight:600;
            }
            QPushButton:hover { background:#e2e8f0; }
            QPushButton:pressed { background:#c7d2fe; }
        """)
        return b

    def _vline(self):
        l = QFrame()
        l.setFrameShape(QFrame.Shape.VLine)
        l.setStyleSheet("color:#cbd5f5;")
        return l

    def export_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export Pipe Tally", "", "Excel (*.xlsx)")
        if not path:
            return
        self.parent.get_filtered_dataframe().to_excel(path, index=False)
        QMessageBox.information(self, "Exported", "Pipe tally exported successfully!")


# ================= MAIN VIEWER =================

class PipeTallyViewer(QMainWindow):
    def __init__(self, df, parent=None):
        super().__init__(parent)   # ✅ now Qt owns this window


        self.df = df
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)
        self.EXCEL_PATH = self.parent().pipetally_dir

        self.setWindowTitle("Pipe Tally Viewer")
        self.resize(1500, 820)

        central = QWidget()
        self.setCentralWidget(central)

        layout = QVBoxLayout(central)
        layout.setSpacing(0)
        layout.setContentsMargins(0,0,0,0)

        # Toolbar
        self.tool_panel = PipeTallyToolPanel(self)
        layout.addWidget(self.tool_panel)

        # Search (hidden initially)
        self.search = QLineEdit()
        self.search.setPlaceholderText("Search Pipe Tally...")
        self.search.setFixedHeight(38)
        self.search.setStyleSheet("margin:6px 14px;")
        self.search.setVisible(False)
        layout.addWidget(self.search)

        # Table
        self.table = QTableView()
        layout.addWidget(self.table)

        # -------- Build Model --------
        model = QStandardItemModel()
        model.setRowCount(len(df))
        model.setColumnCount(len(df.columns))
        model.setHorizontalHeaderLabels(df.columns.astype(str).tolist())

        for r in range(len(df)):
            for c, col in enumerate(df.columns):
                val = df.iloc[r, c]

                if col == "Pipe Number" and not pd.isna(val):
                    item = QStandardItem()
                    item.setData(int(val), Qt.ItemDataRole.DisplayRole)
                elif pd.isna(val):
                    item = QStandardItem("")
                elif isinstance(val, (int, float)):
                    item = QStandardItem()
                    item.setData(float(val), Qt.ItemDataRole.DisplayRole)
                else:
                    item = QStandardItem(str(val))

                item.setEditable(False)
                model.setItem(r, c, item)

        # -------- Sorting + Filtering --------
        self.proxy = QSortFilterProxyModel()
        self.proxy.setSourceModel(model)
        self.proxy.setSortRole(Qt.ItemDataRole.DisplayRole)
        self.proxy.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)

        self.table.setModel(self.proxy)
        self.table.setSortingEnabled(True)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setStretchLastSection(True)

        self.table.sortByColumn(
            df.columns.get_loc("Pipe Number"),
            Qt.SortOrder.AscendingOrder
        )

        self.search.textChanged.connect(self.proxy.setFilterFixedString)
        self.fit_columns_to_headers()

        # -------- Light Theme --------
        self.setStyleSheet("""
            QMainWindow { background:#f8fafc; }
            QTableView {
                background:white;
                gridline-color:#e2e8f0;
                font-size:13px;
                alternate-background-color:#f1f5f9;
                selection-background-color:#2563eb;
                selection-color:white;
            }
            QHeaderView::section {
                background:#e2e8f0;
                padding:7px;
                border:0;
                font-weight:600;
            }
            QLineEdit {
                background:white;
                border:1px solid #cbd5f5;
                border-radius:6px;
                padding:7px;
                font-size:13px;
            }
        """)
    def showEvent(self, event):
        super().showEvent(event)
        if hasattr(self.parent(), "_pipe_spinner") and self.parent()._pipe_spinner:
            self.parent()._pipe_spinner.close()
            self.parent()._pipe_spinner = None



    def toggle_search_bar(self):
        self.search.setVisible(not self.search.isVisible())
        if self.search.isVisible():
            self.search.setFocus()

    def fit_columns_to_headers(self):
        fm = self.table.fontMetrics()
        for i in range(self.proxy.columnCount()):
            header_text = self.proxy.headerData(i, Qt.Orientation.Horizontal)
            w = fm.horizontalAdvance(str(header_text)) + 28
            self.table.setColumnWidth(i, w)
            self.table.resizeColumnToContents(i)

    def get_filtered_dataframe(self):
        rows = []
        for r in range(self.proxy.rowCount()):
            row = {}
            for c in range(self.proxy.columnCount()):
                idx = self.proxy.index(r, c)
                row[self.df.columns[c]] = idx.data()
            rows.append(row)
        return pd.DataFrame(rows)

    def reload_from_excel(self):
        print("🔄 Reloading pipe tally...")

        df = pd.read_excel(self.parent().pipetally_dir)
        self.df = df

        model = QStandardItemModel()
        model.setRowCount(len(df))
        model.setColumnCount(len(df.columns))
        model.setHorizontalHeaderLabels(df.columns.astype(str).tolist())

        for r in range(len(df)):
            for c, col in enumerate(df.columns):
                val = df.iloc[r, c]

                if col == "Pipe Number" and not pd.isna(val):
                    item = QStandardItem()
                    item.setData(int(val), Qt.ItemDataRole.DisplayRole)
                elif pd.isna(val):
                    item = QStandardItem("")
                elif isinstance(val, (int, float)):
                    item = QStandardItem()
                    item.setData(float(val), Qt.ItemDataRole.DisplayRole)
                else:
                    item = QStandardItem(str(val))

                item.setEditable(False)
                model.setItem(r, c, item)

        self.proxy.setSourceModel(model)

        # Keep original sorting behavior
        self.table.sortByColumn(
            df.columns.get_loc("Pipe Number"),
            Qt.SortOrder.AscendingOrder
        )

        self.fit_columns_to_headers()
        print("✔ Table refreshed identically")

# ================= PIPE TALLY LAUNCHER =================





class PipeTallySpinner(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(260, 110)
        self.setWindowTitle("Loading Pipe Tally")
        self.setModal(True)

        v = QVBoxLayout(self)
        v.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.lbl = QLabel("Loading Pipe Tally… 0%")
        self.lbl.setStyleSheet("font-size:13px;font-weight:600;")

        self.bar = QProgressBar()
        self.bar.setRange(0,100)

        v.addWidget(self.lbl)
        v.addWidget(self.bar)

    def set_progress(self, v):
        self.bar.setValue(v)
        self.lbl.setText(f"Loading Pipe Tally… {v}%")


class PipeTallyExcelLoader(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(object)

    def __init__(self, path):
        super().__init__()
        self.path = path

    def run(self):
        import openpyxl
        wb = openpyxl.load_workbook(self.path, read_only=True)
        ws = wb.active
        total = ws.max_row - 1
        rows = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            rows.append(row)
            if i % 500 == 0:
                self.progress.emit(int((i/total)*100))

        import pandas as pd
        df = pd.DataFrame(rows, columns=[c.value for c in ws[1]])
        self.progress.emit(100)
        self.finished.emit(df)


def open_pipetally(self, excel_path=None):
    print(f"inside open pipetally path: {self.pipetally_dir}")
    if excel_path is None:
        excel_path = self.pipetally_dir

    if hasattr(self, "_pipe_tally_instance") and self._pipe_tally_instance:
        self._pipe_tally_instance.close()

    self._pipe_spinner = PipeTallySpinner(self)
    self._pipe_spinner.show()

    self._pipe_loader = PipeTallyExcelLoader(excel_path)
    self._pipe_loader.progress.connect(self._pipe_spinner.set_progress)
    self._pipe_loader.finished.connect(lambda df: _finish_pipe_tally_load(self, df))
    self._pipe_loader.start()




def _finish_pipe_tally_load(self, df):
    # DO NOT close spinner here anymore

    self._pipe_tally_instance = PipeTallyViewer(df, parent=self)
    self._pipe_tally_instance.show()
    self._pipe_tally_instance.raise_()
    self._pipe_tally_instance.activateWindow()



