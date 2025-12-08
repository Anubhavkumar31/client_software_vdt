import sys
import os
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QToolButton,
    QMenu, QStackedLayout, QPushButton, QComboBox, QTableWidget,
    QTableWidgetItem, QFileDialog, QMessageBox, QSplitter, QLabel
)
from PyQt6.QtGui import QColor, QBrush, QFont, QAction
from PyQt6.QtCore import Qt



class PipeTallyComparisonUI(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self.left_df = None
        self.right_df = None
        self.initUI()

    def parse_orientation(self, value):
        from datetime import datetime
        try:
            return datetime.strptime(str(value).strip(), '%H:%M:%S')
        except:
            return None

    def initUI(self):
        self.main_layout = QVBoxLayout(self)

        self.compare_button = QPushButton("Compare Tally Files")
        # self.compare_button.setStyleSheet("padding: 8px; font-weight: bold; background-color: #4CAF50; color: white; border-radius: 5px;")
        self.compare_button.setEnabled(False)
        self.compare_button.clicked.connect(self.compare_data)
        self.main_layout.addWidget(self.compare_button)

        load_layout = QHBoxLayout()
        left_layout = QVBoxLayout()
        right_layout = QVBoxLayout()

        self.load_left_button = QPushButton('Load Pipe Tally 1')
        self.load_left_button.clicked.connect(self.load_left_data)
        left_layout.addWidget(self.load_left_button)

        self.left_column_combo = QComboBox()
        self.left_column_combo.addItem("All")
        left_layout.addWidget(self.left_column_combo)

        self.left_table = QTableWidget()
        left_layout.addWidget(self.left_table)

        self.load_right_button = QPushButton('Load Pipe Tally 2')
        self.load_right_button.clicked.connect(self.load_right_data)
        right_layout.addWidget(self.load_right_button)

        self.right_column_combo = QComboBox()
        self.right_column_combo.addItem("All")
        right_layout.addWidget(self.right_column_combo)

        self.right_table = QTableWidget()
        right_layout.addWidget(self.right_table)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        left_widget = QWidget()
        left_widget.setLayout(left_layout)

        right_widget = QWidget()
        right_widget.setLayout(right_layout)

        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)

        self.main_layout.addLayout(load_layout)
        self.main_layout.addWidget(splitter)
        self.setLayout(self.main_layout)

    def load_left_data(self):
        file_name, _ = QFileDialog.getOpenFileName(self, 'Open Pipe Tally 1 File', '',
                                                   'Excel Files (*.xlsx);;All Files (*)')
        if file_name:
            try:
                df = pd.read_excel(file_name, header=0)
                df.columns = df.columns.str.strip().str.replace('’', "'").str.replace("  ", " ")
                df = df.dropna(how='all')
                self.left_df = df
                self.update_table(self.left_table, df)
                self.left_column_combo.clear()
                self.left_column_combo.addItem("All")
                self.left_column_combo.addItems(df.columns.tolist())
                self.check_ready_to_compare()
            except Exception as e:
                self.show_error_message(f"Failed to load data from Pipe Tally 1: {e}")

    def load_right_data(self):
        file_name, _ = QFileDialog.getOpenFileName(self, 'Open Pipe Tally 2 File', '',
                                                   'Excel Files (*.xlsx);;All Files (*)')
        if file_name:
            try:
                df = pd.read_excel(file_name, header=0)
                df.columns = df.columns.str.strip().str.replace('’', "'").str.replace("  ", " ")
                df = df.dropna(how='all')
                self.right_df = df
                self.update_table(self.right_table, df)
                self.right_column_combo.clear()
                self.right_column_combo.addItem("All")
                self.right_column_combo.addItems(df.columns.tolist())
                self.check_ready_to_compare()
            except Exception as e:
                self.show_error_message(f"Failed to load data from Pipe Tally 2: {e}")

    def check_ready_to_compare(self):
        if self.left_df is not None and self.right_df is not None:
            self.compare_button.setEnabled(True)

    def compare_data(self):
        try:
            df1 = self.left_df.copy()
            df2 = self.right_df.copy()
            merged = pd.merge(df1, df2, on=['Abs. Distance (m)'], how='outer', suffixes=('_x', '_y'))
            is_duplicate_x = merged.duplicated(
                subset=['Abs. Distance (m)'] + [col for col in merged.columns if col.endswith('_x')], keep=False)

            results = []
            found_first_weld = False
            right_counts = df2['Abs. Distance (m)'].value_counts().to_dict()

            for _, row in merged.iterrows():
                status = 'Unmatched'
                match_details = []

                ori_x = row.get("Orientation O'clock_x")
                ori_y = row.get("Orientation O'clock_y")
                dist_us_x = row.get("Distance to U/S GW, m_x")
                dist_us_y = row.get("Distance to U/S GW, m_y")
                len_x = row.get('Length (mm)_x')
                len_y = row.get('Length (mm)_y')
                wid_x = row.get('Width (mm)_x')
                wid_y = row.get('Width (mm)_y')
                dep_x = row.get('Depth (mm)_x')
                dep_y = row.get('Depth (mm)_y')
                abs_dist = row.get('Abs. Distance (m)')

                if not found_first_weld and pd.notna(row['Pipe Length (m)_x']) and pd.notna(
                        row['Pipe Length (m)_y']) and abs(row['Pipe Length (m)_x'] - row['Pipe Length (m)_y']) <= 50:
                    found_first_weld = True
                    status = 'Matched First Weld'

                if pd.notna(row['Pipe Length (m)_x']) and pd.notna(row['Pipe Length (m)_y']) and abs(
                        row['Pipe Length (m)_x'] - row['Pipe Length (m)_y']) <= 50:
                    if row['Feature Identification_x'] == row['Feature Identification_y']:
                        feature_type = str(row['Feature Identification_x']).strip().lower()
                        if 'corrosion' in feature_type:
                            parsed_ori_x = self.parse_orientation(ori_x)
                            parsed_ori_y = self.parse_orientation(ori_y)
                            if parsed_ori_x and parsed_ori_y and abs(
                                    (parsed_ori_x - parsed_ori_y).total_seconds()) <= 1800:
                                match_details.append("Orientation")
                            if pd.notna(dist_us_x) and pd.notna(dist_us_y) and abs(dist_us_x - dist_us_y) <= 50:
                                match_details.append("Upstream Distance")
                            if pd.notna(len_x) and pd.notna(len_y) and abs(len_x - len_y) <= 50:
                                match_details.append("Length")
                            if pd.notna(wid_x) and pd.notna(wid_y) and abs(wid_x - wid_y) <= 50:
                                match_details.append("Width")
                            if pd.notna(dep_x) and pd.notna(dep_y) and abs(dep_x - dep_y) <= 50:
                                match_details.append("Depth")

                            required = ["Orientation", "Upstream Distance", "Length", "Width", "Depth"]
                            if all(f in match_details for f in required):
                                status = 'Full Matched'
                            else:
                                not_matched = [f for f in required if f not in match_details]
                                status = 'Not Matched (Corrosion + Metal Loss): ' + ', '.join(not_matched)
                        else:
                            status = 'Matched (Other Feature Type , Not corrosion)'
                    else:
                        status = 'Not Matched (Feature Types are Different)'

                row_dict = row.to_dict()
                row_dict['Status'] = status

                if (
                        not status.startswith("Full Matched") and
                        is_duplicate_x.loc[_] and (
                        row.get("Feature Identification_x") != row.get("Feature Identification_y") or
                        row.get("Orientation O'clock_x") != row.get("Orientation O'clock_y")
                )
                ):
                    for col in list(row_dict.keys()):
                        if col.endswith('_x'):
                            row_dict[col] = "nan"

                results.append(row_dict)

            result_df = pd.DataFrame(results)
            self.main_window.show_comparison_result(result_df)

        except Exception as e:
            self.show_error_message(f"Error during comparison: {e}")

    def update_table(self, table, df):
        table.setRowCount(df.shape[0])
        table.setColumnCount(df.shape[1])
        table.setHorizontalHeaderLabels(df.columns.astype(str))

        for row in range(df.shape[0]):
            for col in range(df.shape[1]):
                item = QTableWidgetItem(str(df.iat[row, col]))
                # Align flags updated for PyQt6
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
                # Clear editable flag (use bitwise AND NOT with ItemFlag)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                item.setToolTip(str(df.columns[col]))
                status = str(df.at[row, 'Status']) if 'Status' in df.columns else ''
                if status.startswith('Full Matched'):
                    item.setBackground(QBrush(QColor(173, 216, 230)))  # light blue RGB
                elif status.startswith('Not Matched'):
                    item.setBackground(QBrush(QColor(255, 99, 71)))  # tomato
                elif status.startswith('Matched First Weld'):
                    item.setBackground(QBrush(QColor(200, 255, 200)))  # light green
                table.setItem(row, col, item)

        table.resizeColumnsToContents()
        table.resizeRowsToContents()

    def show_error_message(self, message):
        QMessageBox.critical(self, "Error", message)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Dashboard")
        self.setGeometry(100, 100, 1000, 600)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        top_bar = QHBoxLayout()
        self.stack_layout = QStackedLayout()

        self.menu_button = QToolButton()
        self.menu_button.setText("\u2630")
        self.menu_button.setStyleSheet("font-size: 20px; padding: 5px 5px;")
        # set PopupMode updated for PyQt6
        self.menu_button.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)

        menu = QMenu()
        pipe_tally_action = QAction("Pipe Tally Comparison", self)
        pipe_tally_action.triggered.connect(self.show_pipe_tally)
        menu.addAction(pipe_tally_action)

        manual_action = QAction("User Manual", self)
        manual_action.triggered.connect(self.open_manual)
        menu.addAction(manual_action)

        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)
        menu.addAction(exit_action)

        self.menu_button.setMenu(menu)
        top_bar.addWidget(self.menu_button)

        self.welcome_label = QLabel("Welcome to Pipe Tally App")
        self.welcome_label.setStyleSheet("font-size: 16px; color: gray;")
        top_bar.addWidget(self.welcome_label)
        top_bar.addStretch()

        self.default_page = QWidget()
        self.stack_layout.addWidget(self.default_page)

        self.pipe_tally_ui = PipeTallyComparisonUI(self)
        self.stack_layout.addWidget(self.pipe_tally_ui)

        self.comparison_table = QTableWidget()
        self.comparison_page = QWidget()
        layout = QVBoxLayout()

        header_layout = QHBoxLayout()
        header_label = QLabel("Comparison Result")
        header_layout.addWidget(header_label)
        header_layout.addStretch()

        self.full_matched_button = QPushButton("Full Matched Results")
        self.full_matched_button.setStyleSheet(
            "padding: 12px 24px; background-color: #2196F3; color: white; border: none; border-radius: 4px; font-weight: bold;font-size: 12px;")
        self.full_matched_button.clicked.connect(self.show_full_matched_results)
        header_layout.addWidget(self.full_matched_button)

        layout.addLayout(header_layout)
        layout.addWidget(self.comparison_table)
        self.comparison_page.setLayout(layout)
        self.stack_layout.addWidget(self.comparison_page)

        # Add full matched page
        self.full_matched_table = QTableWidget()
        self.full_matched_page = QWidget()
        full_layout = QVBoxLayout()

        full_header_layout = QHBoxLayout()
        full_header_label = QLabel("Full Matched Results")
        full_header_layout.addWidget(full_header_label)
        full_header_layout.addStretch()

        back_button = QPushButton("Back to Comparison")
        back_button.setStyleSheet(
            "padding: 12px 24px; background-color: #f44336; color: white; border: none; border-radius: 4px;font-weight: bold;font-size: 12px;")
        back_button.clicked.connect(lambda: self.stack_layout.setCurrentWidget(self.comparison_page))
        full_header_layout.addWidget(back_button)

        full_layout.addLayout(full_header_layout)
        full_layout.addWidget(self.full_matched_table)
        self.full_matched_page.setLayout(full_layout)
        self.stack_layout.addWidget(self.full_matched_page)
        download_button = QPushButton("Download")
        download_button.setStyleSheet(
            "padding: 12px 24px; background-color: #4CAF50; color: white; border: none; border-radius: 4px; font-weight: bold;font-size: 12px;")
        download_button.clicked.connect(self.download_full_matched)
        full_header_layout.addWidget(download_button)

        main_layout = QVBoxLayout()
        main_layout.addLayout(top_bar)
        main_layout.addLayout(self.stack_layout)
        central_widget.setLayout(main_layout)

    def show_pipe_tally(self):
        self.welcome_label.hide()
        self.stack_layout.setCurrentWidget(self.pipe_tally_ui)

    def get_manual_path(self):
        # Automatically detects folder of the currently running script
        base_path = os.path.dirname(os.path.abspath(__file__))

        # Manual must be placed next to UI_compare.py
        manual_path = os.path.join(base_path, "ffp_manual.pdf")
        return manual_path

    def open_manual(self):
        manual_path = self.get_manual_path()

        if os.path.exists(manual_path):
            os.startfile(manual_path)  # Opens with default PDF viewer
        else:
            QMessageBox.warning(
                self,
                "Manual Missing",
                f"FFP manual not found at:\n{manual_path}\n\n"
                "Place ffp_manual.pdf in the same folder as UI_compare.py."
            )

    def show_comparison_result(self, df):
        try:
            # Insert a visual separator column after the last _x column
            cols = df.columns.tolist()
            x_indices = [i for i, col in enumerate(cols) if col.endswith('_x')]
            separator_index = max(x_indices) + 1 if x_indices else len(cols)

            cols.insert(separator_index, '---')  # separator label
            df.insert(separator_index, '---', ['' for _ in range(df.shape[0])])

            self.comparison_table.setRowCount(df.shape[0])
            self.comparison_table.setColumnCount(len(cols))
            self.comparison_table.setHorizontalHeaderLabels(cols)
            font_bold = QFont()
            font_bold.setBold(True)
            for i, header in enumerate(cols):
                item = self.comparison_table.horizontalHeaderItem(i)
                if item and header.strip() == "Abs. Distance (m)":
                    item.setFont(font_bold)

            for row in range(df.shape[0]):
                for col, column_name in enumerate(cols):
                    item = QTableWidgetItem(str(df.iat[row, col]))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

                    if column_name == '---':
                        item.setBackground(QBrush(QColor(211, 211, 211)))  # separator grey RGB
                    else:
                        status = str(df.at[row, 'Status']) if 'Status' in df.columns else ''
                        if status.startswith('Full Matched'):
                            item.setBackground(QBrush(QColor(173, 216, 230)))  # light blue RGB
                        elif status.startswith('Not Matched'):
                            item.setBackground(QBrush(QColor(255, 99, 71)))  # tomato
                        elif status.startswith('Matched First Weld'):
                            item.setBackground(QBrush(QColor(200, 255, 200)))  # light green

                    self.comparison_table.setItem(row, col, item)

            self.comparison_table.resizeColumnsToContents()
            self.comparison_table.resizeRowsToContents()
            self.stack_layout.setCurrentWidget(self.comparison_page)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to show comparison results: {e}")

    def show_full_matched_results(self):
        try:
            original_headers = [self.comparison_table.horizontalHeaderItem(i).text() for i in
                                range(self.comparison_table.columnCount())]
            separator_index = max(i for i, col in enumerate(original_headers) if col.endswith('_x')) + 1
            modified_headers = original_headers[:]
            if '---' not in modified_headers:
                modified_headers.insert(separator_index, '---')

            full_matched_data = []

            for row in range(self.comparison_table.rowCount()):
                status_item = self.comparison_table.item(row, original_headers.index("Status"))
                if status_item and status_item.text().startswith("Full Matched"):
                    row_data = []
                    for col in range(len(original_headers)):
                        item = self.comparison_table.item(row, col)
                        row_data.append(item.text() if item else '')
                    if '---' not in original_headers:
                        row_data.insert(separator_index, '')  # blank for separator
                    full_matched_data.append(row_data)

            if not full_matched_data:
                QMessageBox.information(self, "No Data", "No 'Full Matched' rows found.")
                return

            self.full_matched_table.setRowCount(len(full_matched_data))
            self.full_matched_table.setColumnCount(len(modified_headers))
            self.full_matched_table.setHorizontalHeaderLabels(modified_headers)
            font_bold = QFont()
            font_bold.setBold(True)
            for i, header in enumerate(modified_headers):
                item = self.full_matched_table.horizontalHeaderItem(i)
                if item and header.strip() == "Abs. Distance (m)":
                    item.setFont(font_bold)

            for row_idx, row_data in enumerate(full_matched_data):
                for col_idx, value in enumerate(row_data):
                    item = QTableWidgetItem(value)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    if modified_headers[col_idx] == '---':
                        item.setBackground(QBrush(QColor(211, 211, 211)))  # separator grey
                    self.full_matched_table.setItem(row_idx, col_idx, item)

            self.full_matched_table.resizeColumnsToContents()
            self.full_matched_table.resizeRowsToContents()
            self.stack_layout.setCurrentWidget(self.full_matched_page)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to show full matched results: {e}")

    def download_full_matched(self):
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill

            headers = [self.full_matched_table.horizontalHeaderItem(i).text() for i in
                       range(self.full_matched_table.columnCount())]
            data = []

            for row in range(self.full_matched_table.rowCount()):
                row_data = []
                for col in range(self.full_matched_table.columnCount()):
                    item = self.full_matched_table.item(row, col)
                    row_data.append(item.text() if item else '')
                data.append(row_data)

            if not data:
                QMessageBox.warning(self, "Warning", "No data to download.")
                return

            df = pd.DataFrame(data, columns=headers)
            file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "full_matched_results.xlsx",
                                                       "Excel Files (*.xlsx)")
            if file_path:
                df.to_excel(file_path, index=False)

                # Load with openpyxl for formatting
                wb = load_workbook(file_path)
                ws = wb.active

                yellow_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')  # Light yellow hex (openpyxl ok)
                blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light blue hex
                bold_font = Font(bold=True)

                for col_index, header in enumerate(headers, start=1):
                    col_letter = ws.cell(row=1, column=col_index).column_letter

                    # Bold header
                    header_cell = ws.cell(row=1, column=col_index)
                    header_cell.font = bold_font

                    # Fill color
                    if header.endswith('_x'):
                        fill = yellow_fill
                    elif header.endswith('_y'):
                        fill = blue_fill
                    else:
                        fill = None

                    if fill:
                        for row in range(2, ws.max_row + 1):
                            ws.cell(row=row, column=col_index).fill = fill

                    # Auto column width
                    max_length = max(
                        len(str(ws.cell(row=row, column=col_index).value or ''))
                        for row in range(1, ws.max_row + 1)
                    )
                    ws.column_dimensions[col_letter].width = max_length + 2  # +2 for padding

                wb.save(file_path)
                QMessageBox.information(self, "Success", f"File saved to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save file: {e}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    main_win = MainWindow()
    main_win.show()
    sys.exit(app.exec())
